# -*- coding: utf-8 -*-
"""
Created on Wed May 20 16:16:35 2020

@author: brody
Code to grab data from Daily Treasury Statement extract 2 items of interest and reformat into excel and then csv which feeds a power bi 
dashboard.  The dash updates automatically from one drive once the csv is updated.  The script runs automatically using windows task 
scheduler so the dashbaord is automatically updated daily.  On weekends there is no new data so the code will not find a file and exit. 
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import date, timedelta
import holidays
import sys

# find todays date and dict of holidays
today = date.today() 
hol = holidays.UnitedStates(years= today.year).keys()
day = today.strftime('%A') # get day of week for today

# find if holiday with in last 2  bussiness days which delays release of treasury data
if today in hol:
    sys.exit()  # no new data available today
elif today - timedelta(days=3) in hol and day == 'Monday':
    dayhol = (today - timedelta(days=3)).strftime('%A')# day of the week of the holiday
    delay = 1
elif today - timedelta(days=4) in hol and day == 'Tuesday':
    dayhol = (today - timedelta(days=4)).strftime('%A')
    delay = 1
elif today - timedelta(days=4) in hol and day == 'Monday':
    dayhol = (today - timedelta(days=4)).strftime('%A')
    delay = 1
elif today - timedelta(days=1) in hol: # yesterday was holiday
    dayhol = (today - timedelta(days=1)).strftime('%A') 
    delay = 1 
elif today - timedelta(days=2) in hol:
    dayhol = (today - timedelta(days=2)).strftime('%A')
    delay = 1
else:
    delay = 0 

# on weekends 
if not np.is_busday(today):
    sys.exit()  # no new data available today

# if no recent holidays calculate which days treasury data is availble on Mon and Tues report is from before the weekend
if delay == 0:
    if day == 'Monday' or day == "Tuesday":
        Target = today - timedelta(days=4)
    else:
        Target = today - timedelta(days=2)
# if there was a recent holiday there is a delay in which report is availble depending on how weekends fall out
else:
    if dayhol ==  'Monday' or dayhol == "Friday":
        Target = today - timedelta(days=5)
    elif dayhol == "Tuesday":
        if day == 'Wednesday':
            Target = today - timedelta(days=5)
        else:
            Target = today - timedelta(days=3)
    elif dayhol == 'Wednesday':
        Target = today - timedelta(days=3)
    else:
        if day == 'Friday':
            Target = today - timedelta(days=3)
        else:
            Target = today - timedelta(days=5)

targ = Target.strftime("%m/%d/%y")

# figure out the url extension for the new treasury report availble based on the date
yr = targ.split('/')[2]
d = targ.split('/')[1]
mon = targ.split('/')[0]
ext = yr+mon+d+'00.xlsx'
head = 'https://fsapps.fiscal.treasury.gov/dts/files/'
url = head+ext

# grab daily report to add to the records
try:
    df = pd.read_excel(url)
except:
    sys.exit()

# extract the date
d = df.columns[0]


# drop empty column that appears on some uploads of the data set
typ = 'A'
if df.columns[-1] == 'Unnamed: 6':
    df.drop('Unnamed: 6', axis=1, inplace=True)
    typ = 'B'

# rename the columns
df.columns = ['Table','Account', 'Closing today', 'Today', 'Month to date', 'FYTD' ]

#Track Fed MLF loans

# find entry for MLF seems to change every time one time b was 73
MLF = df[df['Account']=='ESF - Economic Recovery Programs']


data = pd.DataFrame()

data = data.append(MLF)



data['Date'] = d
data.columns = [['Table','Account', 'Closing today', 'Today', 'Month to date', 'FYTD','Date']]



# code was adopted from here: https://medium.com/better-programming/using-python-pandas-with-excel-d5082102ca27

writer = pd.ExcelWriter('ESF.xlsx', engine='openpyxl')
# try to open an existing workbook
writer.book = load_workbook('ESF.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'ESF.xlsx')
# write out the new sheet
data.to_excel(writer,index=False,header=False,startrow=len(reader)+1)

writer.close()

### Track witholdings from income and employment tax

# find entry for withholdings 

WH = df[df['Account'] =='Withheld Income and Employment Taxes']


# same as code above
data = pd.DataFrame()

data = data.append(WH)
data['Date'] = d

data.columns = [['Table','Account', 'Closing today', 'Today', 'Month to date', 'FYTD','Date']]



writer = pd.ExcelWriter('Withholdings.xlsx', engine='openpyxl')
# try to open an existing workbook
writer.book = load_workbook('Withholdings.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'Withholdings.xlsx')
# write out the new sheet
data.to_excel(writer,index=False,header=False,startrow=len(reader)+1)

writer.close()

#pull up excel and sort by date
cs = pd.read_excel('Withholdings.xlsx')

# in pandas to convert a column of strings into date object
cs['Date'] = pd.to_datetime(cs['Date'])

# sort
cs = cs.sort_values('Date')

# these numbers are in millions better for dashboard to convert to millions
def convert(num):
    if num < 1000000:
        res = num * 1000000
        return res
    else:
        return num
cs['Today'] = cs['Today'].apply(convert)
cs['Month to date'] = cs['Month to date'].apply(convert)


cs.to_csv('Withheld.csv')

df = pd.read_excel('ESF.xlsx')

# convert strings into date object
df['Date'] = pd.to_datetime(df['Date'])

# sort
df = df.sort_values('Date')

# these numbers are in millions better for dashboard to convert to millions
def convert(num):
    if num < 1000000:
        res = num * 1000000
        return res
    else:
        return num
df['Today'] = df['Today'].apply(convert)
df['Month to date'] = df['Month to date'].apply(convert)
df['FYTD'] = df['FYTD'].apply(convert)


df.to_csv('ESF.csv')
